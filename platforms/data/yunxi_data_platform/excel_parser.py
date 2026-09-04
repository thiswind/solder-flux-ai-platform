"""Excel parsing module for the data management platform.

Extracted from the original ``excel/data_column.py`` script and integrated
into the ``yunxi_data_platform`` package so the pipeline no longer depends on
files living outside the package directory.
"""

from __future__ import annotations

import os
import re

import numpy as np
import openpyxl
import pandas as pd

# ---------------------------------------------------------------------------
# Patterns
# ---------------------------------------------------------------------------

MERGE_KEY_SPLIT_PATTERN = re.compile(r"[/,，、;；]+")
CHEMICAL_ELEMENT_PATTERN = re.compile(r"^[A-Za-z]{1,3}$")
PARTICLE_LABEL_PATTERN = re.compile(
    r"(<\s*\d+\s*[µμu]m|\d+\s*[~～-]\s*\d+\s*[µμu]m|>\s*\d+\s*[µμu]m)",
    re.IGNORECASE,
)

# ---------------------------------------------------------------------------
# Text utilities
# ---------------------------------------------------------------------------


def normalize_text(value):
    return str(value).replace("\n", " ").strip()


def split_batch_keys(key_str):
    if pd.isna(key_str):
        return []
    return [part.strip() for part in MERGE_KEY_SPLIT_PATTERN.split(str(key_str)) if part.strip()]


def normalize_particle_label(value):
    text = normalize_text(value)
    if not text or text.lower() == "nan":
        return None

    text = text.replace("μ", "µ").replace("um", "µm").replace("UM", "µm")
    match = PARTICLE_LABEL_PATTERN.search(text)
    if not match:
        return None

    label = match.group(1)
    label = re.sub(r"\s+", "", label)
    label = label.replace("μ", "µ").replace("u", "µ").replace("U", "µ")
    label = label.replace("~", "～")
    label = re.sub(r"(\d)-(\d)", r"\1～\2", label)
    return label


def particle_label_sort_key(column_name):
    label = column_name.split("_", 2)[-1]
    label = normalize_particle_label(label) or label

    less_match = re.match(r"<(\d+)µm", label)
    if less_match:
        return (0, int(less_match.group(1)), int(less_match.group(1)))

    range_match = re.match(r"(\d+)～(\d+)µm", label)
    if range_match:
        return (1, int(range_match.group(1)), int(range_match.group(2)))

    greater_match = re.match(r">(\d+)µm", label)
    if greater_match:
        return (2, int(greater_match.group(1)), int(greater_match.group(1)))

    return (3, label, label)


# ---------------------------------------------------------------------------
# Cell / sheet extraction
# ---------------------------------------------------------------------------


def find_cell_coordinates(df, keyword):
    for r in range(len(df)):
        for c in range(len(df.columns)):
            val = str(df.iloc[r, c])
            if keyword in val:
                return r, c
    return None


def extract_particle_distribution(df, particle_coord, result):
    r, c = particle_coord
    label_row = None
    particle_cols = []

    for candidate_row in range(r, min(len(df), r + 4)):
        current_cols = []
        seen_labels = set()
        start_col = max(0, c - 1)
        end_col = min(len(df.columns), c + 10)
        for col in range(start_col, end_col):
            label = normalize_particle_label(df.iloc[candidate_row, col])
            if label and label not in seen_labels:
                current_cols.append((col, label))
                seen_labels.add(label)

        if len(current_cols) >= 4:
            label_row = candidate_row
            particle_cols = current_cols[:4]
            break

    if not particle_cols:
        return

    standard_row = label_row + 1
    measured_row = None

    for candidate_row in range(label_row + 1, min(len(df), label_row + 5)):
        row_text = " ".join(
            normalize_text(df.iloc[candidate_row, col]).lower()
            for col in range(min(3, len(df.columns)))
        )
        if "实测值" in row_text or "measured value" in row_text:
            measured_row = candidate_row
            break

    if measured_row is None:
        measured_row = standard_row + 1

    for col, label in particle_cols:
        if standard_row < len(df):
            result[f"粒度分布_标准值_{label}"] = df.iloc[standard_row, col]
        if measured_row < len(df):
            result[f"粒度分布_实测值_{label}"] = df.iloc[measured_row, col]


# ---------------------------------------------------------------------------
# Overall data reader
# ---------------------------------------------------------------------------


def read_overall_data(directory):
    """Read all Excel files in *directory* (recursive) and return a combined DataFrame."""
    print(f"Reading overall data from: {directory}")
    if not os.path.exists(directory):
        print(f"Directory not found: {directory}")
        return pd.DataFrame()

    files = []
    for root, _, filenames in os.walk(directory):
        for filename in filenames:
            if (filename.endswith(".xlsx") or filename.endswith(".xls")) and not filename.startswith("~$"):
                files.append(os.path.join(root, filename))

    return read_overall_data_files(files)


def read_overall_data_files(file_paths):
    """Read a list of overall-data Excel files and return a combined DataFrame.

    Shared by :func:`read_overall_data` (directory scan) and upstream callers that
    supply a content-hash de-duplicated file list.
    """
    print(f"Reading overall data from {len(file_paths)} file(s)")
    all_dfs = []

    for file_path in file_paths:
        print(f"  Loading {os.path.basename(file_path)}...")
        try:
            try:
                wb = openpyxl.load_workbook(file_path, data_only=True)
            except Exception as e:
                print(f"    Failed to load with openpyxl: {e}")
                xls = pd.ExcelFile(file_path)
                wb = None

            sheet_names = wb.sheetnames if wb else xls.sheet_names

            sheet_groups = {}
            for sheet in sheet_names:
                base_name = sheet.strip()
                clean_base = re.sub(r"[\(（]?\s*原始\s*[\)）]?", "", base_name).strip()
                if clean_base not in sheet_groups:
                    sheet_groups[clean_base] = []
                sheet_groups[clean_base].append(sheet)

            final_sheets_to_read = []
            for base, sheets in sheet_groups.items():
                if len(sheets) > 1:
                    original_sheet = next((s for s in sheets if "原始" in s), None)
                    if original_sheet:
                        final_sheets_to_read.append(original_sheet)
                    else:
                        final_sheets_to_read.append(sheets[0])
                else:
                    final_sheets_to_read.append(sheets[0])

            for sheet_name in final_sheets_to_read:
                clean_sheet_name = sheet_name.strip()
                if clean_sheet_name in ["模板", "清单"]:
                    continue
                if re.match(r"^Sheet\d+$", clean_sheet_name, re.IGNORECASE):
                    continue

                try:
                    df = None
                    if wb:
                        ws = wb[sheet_name]
                        visible_rows = []
                        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                            if ws.row_dimensions[i].hidden:
                                continue
                            visible_rows.append(row)

                        if visible_rows:
                            header = visible_rows[0]
                            data = visible_rows[1:]
                            header = [
                                str(h) if h is not None else f"Unnamed:{i}"
                                for i, h in enumerate(header)
                            ]
                            if len(header) != len(set(header)):
                                seen = {}
                                new_header = []
                                for h in header:
                                    if h in seen:
                                        seen[h] += 1
                                        new_header.append(f"{h}.{seen[h]}")
                                    else:
                                        seen[h] = 0
                                        new_header.append(h)
                                header = new_header
                            df = pd.DataFrame(data, columns=header)
                        else:
                            continue
                    else:
                        df = pd.read_excel(xls, sheet_name=sheet_name)

                    if df is None or df.empty:
                        continue

                    col_map = {}
                    for col in df.columns:
                        col_str = str(col).strip()
                        col_clean = re.sub(r"\s+", "", col_str)

                        if "助焊膏" in col_clean and "型号" in col_clean:
                            col_map[col] = "助焊膏"
                        elif "合金" in col_clean and "比例" in col_clean:
                            col_map[col] = "合金含量（%）"
                        elif "助焊剂" in col_clean and "比例" in col_clean:
                            col_map[col] = "助焊剂比例%"
                        elif "锡粉氧含量" in col_clean:
                            col_map[col] = "锡粉氧含量"
                        elif "黏度初值" in col_clean:
                            col_map[col] = "黏度初值"
                        elif "产品批号" in col_clean:
                            col_map[col] = "产品批号"
                        elif "锡粉批号" in col_clean:
                            col_map[col] = "锡粉批号"

                    df = df.rename(columns=col_map)

                    if "产品批号" in df.columns:
                        df = df.dropna(subset=["产品批号"])
                    else:
                        continue

                    powder_batch_col = None
                    for col in df.columns:
                        if "锡粉批号" in str(col):
                            powder_batch_col = col
                            break

                    if powder_batch_col:
                        df[powder_batch_col] = df[powder_batch_col].astype(str).str.strip()
                        df["merge_key"] = df[powder_batch_col]
                        df["merge_key"] = df["merge_key"].str.replace("，", ",").str.replace("、", ",")
                        df["来源Sheet"] = sheet_name
                        df["来源文件"] = os.path.basename(file_path)
                        all_dfs.append(df)

                except Exception as e:
                    print(f"    Error reading sheet '{sheet_name}': {e}")

        except Exception as e:
            print(f"  Error reading {os.path.basename(file_path)}: {e}")

    if not all_dfs:
        return pd.DataFrame()

    combined_df = pd.concat(all_dfs, ignore_index=True)

    cols_to_check = [c for c in combined_df.columns if c not in ["来源文件", "来源Sheet"]]
    count_before = len(combined_df)
    combined_df.drop_duplicates(subset=cols_to_check, keep="first", inplace=True)
    count_after = len(combined_df)
    print(f"Dropped {count_before - count_after} duplicate rows.")
    print(f"Unique raw records (pre-explosion): {count_after}")

    combined_df = combined_df.dropna(subset=["merge_key"])
    combined_df["merge_key"] = combined_df["merge_key"].astype(str).str.strip()
    combined_df = combined_df[combined_df["merge_key"] != "nan"]
    combined_df = combined_df[combined_df["merge_key"] != ""]

    print(f"Total records in overall data: {len(combined_df)}")
    return combined_df


# ---------------------------------------------------------------------------
# Specific data reader
# ---------------------------------------------------------------------------


def extract_specific_info_from_sheet(df, file_name, sheet_name):
    try:
        result = {
            "来源文件_specific": file_name,
            "来源Sheet_specific": sheet_name,
        }

        # 1. Batch number
        batch_coord = find_cell_coordinates(df, "生产批号")
        batch_no = None

        if batch_coord:
            r, c = batch_coord
            if c + 2 < len(df.columns):
                val = str(df.iloc[r, c + 2]).strip()
                if val and val.lower() != "nan":
                    batch_no = val
            if not batch_no and c + 1 < len(df.columns):
                val = str(df.iloc[r, c + 1]).strip()
                if val and val.lower() != "nan":
                    batch_no = val

        if not batch_no:
            try:
                val = str(df.iloc[7, 2]).strip()
                if val and val.lower() != "nan":
                    batch_no = val
            except Exception:
                pass

        if batch_no:
            result["生产批号"] = batch_no
        else:
            return None

        # 2. Chemical composition
        chem_coord = find_cell_coordinates(df, "Chemical Composition")
        if not chem_coord:
            chem_coord = find_cell_coordinates(df, "化学成分")

        if chem_coord:
            header_r, header_c = chem_coord
            value_r = header_r + 2
            if value_r < len(df):
                for col in range(len(df.columns)):
                    element_name = normalize_text(df.iloc[header_r, col])
                    if not element_name or not CHEMICAL_ELEMENT_PATTERN.match(element_name):
                        continue
                    if element_name.lower() in ("nan", "化学成分", "chemical composition", "标准值", "实测值"):
                        continue
                    result[element_name] = df.iloc[value_r, col]

        # 3. Particle size distribution
        particle_coord = find_cell_coordinates(df, "粒度分布")
        if particle_coord:
            extract_particle_distribution(df, particle_coord, result)

        # 4. Oxygen content
        oxygen_coord = find_cell_coordinates(df, "氧含量")
        if oxygen_coord:
            r, c = oxygen_coord
            val_r1 = str(df.iloc[r + 1, c]).strip() if r + 1 < len(df) else ""
            val_r3 = str(df.iloc[r + 3, c]).strip() if r + 3 < len(df) else ""

            if "合格" in val_r1 or "Pass" in val_r1:
                result["氧含量_实测值"] = val_r1
            elif "合格" in val_r3 or "Pass" in val_r3:
                result["氧含量_实测值"] = val_r3
            else:
                if r + 1 < len(df):
                    result["氧含量_标准值"] = df.iloc[r + 1, c]
                if r + 3 < len(df):
                    result["氧含量_实测值"] = df.iloc[r + 3, c]

        # 5. Sphericity
        shape_coord = find_cell_coordinates(df, "球型度")
        if shape_coord:
            r, c = shape_coord
            val_r1 = str(df.iloc[r + 1, c]).strip() if r + 1 < len(df) else ""

            if "合格" in val_r1 or "Pass" in val_r1:
                result["球型度_实测值"] = val_r1
            else:
                if r + 2 < len(df):
                    result["球型度_标准值"] = df.iloc[r + 2, c]
                if r + 3 < len(df):
                    result["球型度_实测值"] = df.iloc[r + 3, c]

        return result

    except Exception as e:
        print(f"  Error extracting info from {file_name} - {sheet_name}: {e}")
        return None


def read_specific_data_files(file_paths):
    """Read a list of specific-data Excel files and return a combined DataFrame."""
    data_list = []
    for file_path in file_paths:
        file = os.path.basename(file_path)
        if (file.endswith(".xlsx") or file.endswith(".xls")) and not file.startswith("~$"):
            try:
                xls = pd.ExcelFile(file_path)
                for sheet_name in xls.sheet_names:
                    if re.match(r"^Sheet\d+$", sheet_name.strip(), re.IGNORECASE):
                        continue
                    try:
                        df = pd.read_excel(xls, sheet_name=sheet_name, header=None)
                        extracted_data = extract_specific_info_from_sheet(df, file, sheet_name)
                        if extracted_data:
                            data_list.append(extracted_data)
                    except Exception:
                        pass
            except Exception as e:
                print(f"  Error reading specific file {file}: {e}")

    if not data_list:
        return pd.DataFrame()

    specific_df = pd.DataFrame(data_list)
    print(f"Total records extracted from specific data: {len(specific_df)}")
    return specific_df


def read_specific_data(directory):
    """Traverse *directory* recursively and extract info from all Excel files."""
    print(f"Reading specific data from: {directory}")
    files = []
    if os.path.exists(directory):
        for root, _, filenames in os.walk(directory):
            for f in filenames:
                if (f.endswith(".xlsx") or f.endswith(".xls")) and not f.startswith("~$"):
                    files.append(os.path.join(root, f))
    return read_specific_data_files(files)
